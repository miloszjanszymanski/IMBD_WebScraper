import openpyxl

fileName = input("Provide xlsx filename with proper structure, with extension - example.xlsx: ")
workbook = openpyxl.load_workbook(fileName)

sheet = workbook.active

last_row = sheet.max_row

start_value_a = 1

modified_rows = []

for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    value_c = row[2]

    if isinstance(value_c, str):
        value_c = value_c.strip()

    if ',' in str(value_c):
        split_values = str(value_c).split(',')

        for split_value in split_values:
            modified_rows.append([start_value_a, row[1], split_value.strip()])
            start_value_a += 1
    else:
        modified_rows.append([start_value_a, row[1], value_c.strip()])
        start_value_a += 1


modified_sheet = workbook.create_sheet(title='Genres')

modified_sheet['A1'] = "Index"
modified_sheet['B1'] = "Name"
modified_sheet['C1'] = "Genre"

for row in modified_rows:
    modified_sheet.append(row)

workbook.save(fileName)

