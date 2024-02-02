import openpyxl

fileName= input("Provide xlsx filename with proper structure, with extension - example.xlsx: ")
workbook = openpyxl.load_workbook(fileName)

sheet = workbook.active

last_row = sheet.max_row

start_value_a = 1

modified_rows = []


modified_sheet = workbook.create_sheet(title='Actors')

modified_sheet['A1'] = "Index"
modified_sheet['B1'] = "Name"
modified_sheet['C1'] = "Actor"

for row in modified_rows:
    modified_sheet.append(row)

workbook.save(fileName)
